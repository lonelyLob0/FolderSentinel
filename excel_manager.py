from tomli import load
from folder_manager import FolderManager
from openpyxl import Workbook, load_workbook


class ExcelManager:
    """a"""

    def create_master_workflow(self, folder_path):
        """a"""
        master_directory_path = folder_path + r"\Master"
        master_file_path = master_directory_path + r"\master.xlsx"

        if not FolderManager.directory_exists(master_directory_path):

            fm = FolderManager(folder_path)
            fm.create_directory("Master")

            wb = Workbook()
            wb.active.title = "Summary"
            wb.save(filename=master_file_path)
            wb.close()

    def consolidate_excel_file_in_target(self, origin_path, target_path):
        """a"""
        print(origin_path)
        origin_wb = load_workbook(origin_path)
        target_wb = load_workbook(target_path)

        for sheet in origin_wb.sheetnames:
            origin_current_ws = origin_wb[sheet]
            new_target_ws = target_wb.create_sheet(sheet)

            for row in origin_current_ws:
                for cell in row:
                    new_target_ws[cell.coordinate].value = cell.value

        target_wb.save(target_path)

        origin_wb.close()
        target_wb.close()
